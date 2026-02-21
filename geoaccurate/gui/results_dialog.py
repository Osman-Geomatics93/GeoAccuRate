"""Results display dialogs for GeoAccuRate.

Provides simple QDialog wrappers to show confusion matrix
and per-class metrics in table form, plus an interactive heatmap.
"""

import math
import os

from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtGui import QBrush, QColor
from qgis.PyQt.QtWidgets import (
    QComboBox,
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..domain.models import ConfusionMatrixResult

# Try to import matplotlib Qt backend for the heatmap tab
try:
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
    from ..reporting.chart_renderer import create_extended_heatmap_figure, is_available

    HAS_MPL_QT = is_available()
except Exception:
    HAS_MPL_QT = False

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

_COLORMAPS = [
    "YlOrRd", "YlGn", "Blues", "Greens", "Purples",
    "viridis", "plasma", "coolwarm",
]


class ConfusionMatrixDialog(QDialog):
    """Dialog showing confusion matrix as a table and interactive heatmap."""

    def __init__(self, result: ConfusionMatrixResult, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GeoAccuRate \u2014 Confusion Matrix")
        self.setMinimumSize(700, 550)
        self._result = result
        self._build_ui(result)

    def _build_ui(self, result):
        layout = QVBoxLayout(self)

        if HAS_MPL_QT:
            tabs = QTabWidget()
            tabs.addTab(self._create_table_tab(result), "Table")
            tabs.addTab(self._create_heatmap_tab(result), "Heatmap")
            layout.addWidget(tabs)
        else:
            layout.addWidget(self._create_table_widget(result))

        # Bottom button row
        btn_row = QHBoxLayout()

        if HAS_OPENPYXL:
            btn_excel = QPushButton("Export to Excel...")
            btn_excel.clicked.connect(self._on_export_excel)
            btn_row.addWidget(btn_excel)

        btn_row.addStretch()

        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.close)
        btn_row.addWidget(btn_close)

        layout.addLayout(btn_row)

    # ---- Table tab --------------------------------------------------------

    def _create_table_tab(self, result):
        widget = QWidget()
        vbox = QVBoxLayout(widget)
        vbox.addWidget(self._create_table_widget(result))
        return widget

    def _create_table_widget(self, result):
        labels = [result.class_names.get(lbl, str(lbl)) for lbl in result.class_labels]
        k = len(labels)
        matrix = result.matrix

        table = QTableWidget(k + 1, k + 1)
        table.setHorizontalHeaderLabels(labels + ["Row Total"])
        table.setVerticalHeaderLabels(labels + ["Col Total"])

        diag_color = QBrush(QColor(217, 243, 217))

        for i in range(k):
            for j in range(k):
                item = QTableWidgetItem(str(int(matrix[i, j])))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if i == j:
                    item.setBackground(diag_color)
                table.setItem(i, j, item)

            row_total = QTableWidgetItem(str(int(matrix[i, :].sum())))
            row_total.setTextAlignment(Qt.AlignCenter)
            row_total.setFlags(row_total.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, k, row_total)

        for j in range(k):
            col_total = QTableWidgetItem(str(int(matrix[:, j].sum())))
            col_total.setTextAlignment(Qt.AlignCenter)
            col_total.setFlags(col_total.flags() & ~Qt.ItemIsEditable)
            table.setItem(k, j, col_total)

        grand = QTableWidgetItem(str(int(matrix.sum())))
        grand.setTextAlignment(Qt.AlignCenter)
        grand.setFlags(grand.flags() & ~Qt.ItemIsEditable)
        table.setItem(k, k, grand)

        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        return table

    # ---- Heatmap tab ------------------------------------------------------

    def _create_heatmap_tab(self, result):
        widget = QWidget()
        vbox = QVBoxLayout(widget)

        # Toolbar row
        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("Color scheme:"))

        self._cmap_combo = QComboBox()
        self._cmap_combo.addItems(_COLORMAPS)
        self._cmap_combo.setCurrentIndex(0)
        self._cmap_combo.currentTextChanged.connect(self._on_cmap_changed)
        toolbar.addWidget(self._cmap_combo)

        toolbar.addStretch()

        btn_save = QPushButton("Save Image...")
        btn_save.clicked.connect(self._on_save_image)
        toolbar.addWidget(btn_save)

        vbox.addLayout(toolbar)

        # Canvas
        fig = create_extended_heatmap_figure(result, cmap=_COLORMAPS[0])
        self._canvas = FigureCanvasQTAgg(fig)
        vbox.addWidget(self._canvas, stretch=1)

        return widget

    def _on_cmap_changed(self, cmap_name: str):
        fig = create_extended_heatmap_figure(self._result, cmap=cmap_name)
        self._canvas.figure = fig
        self._canvas.draw()

    def _on_save_image(self):
        path, filt = QFileDialog.getSaveFileName(
            self,
            "Save Heatmap Image",
            "confusion_matrix_heatmap.png",
            "PNG (*.png);;JPEG (*.jpg);;SVG (*.svg);;PDF (*.pdf)",
        )
        if path:
            self._canvas.figure.savefig(path, dpi=300, bbox_inches="tight")

    # ---- Excel export -----------------------------------------------------

    def _on_export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Confusion Matrix to Excel",
            "confusion_matrix.xlsx",
            "Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            _export_confusion_matrix_xlsx(self._result, path)
            QMessageBox.information(self, "Export Complete",
                                   f"Confusion matrix saved to:\n{os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


def _export_confusion_matrix_xlsx(result: ConfusionMatrixResult, path: str):
    """Write the extended confusion matrix to a styled .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confusion Matrix"

    matrix = result.matrix
    k = matrix.shape[0]
    labels = [result.class_names.get(lbl, str(lbl)) for lbl in result.class_labels]

    has_kappa = result.kappa is not None

    # Pre-compute metrics
    row_totals = matrix.sum(axis=1)
    col_totals = matrix.sum(axis=0)
    grand_total = matrix.sum()

    ua_values = []
    for lbl in result.class_labels:
        ua = result.users_accuracy.get(lbl, float("nan"))
        ua_values.append(ua if not math.isnan(ua) else 0.0)

    pa_values = []
    for lbl in result.class_labels:
        pa = result.producers_accuracy.get(lbl, float("nan"))
        pa_values.append(pa if not math.isnan(pa) else 0.0)

    oa = result.overall_accuracy

    # --- Styles ---
    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    num_fmt_int = "0"
    num_fmt_pct = "0.00"

    thin = Side(style="thin")
    thick = Side(style="medium")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="4472C4")
    font_header_white = Font(bold=True, size=11, color="FFFFFF")
    fill_diag = PatternFill("solid", fgColor="D9F3D9")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    fill_metric = PatternFill("solid", fgColor="FFF2CC")
    fill_kappa = PatternFill("solid", fgColor="FCE4D6")

    # Column layout:
    # Col 1: ClassValue | Cols 2..k+1: classes | Col k+2: Total | Col k+3: P_Accuracy | (Col k+4: Kappa)

    # Header row (row 1)
    headers = ["ClassValue"] + labels + ["Total", "P_Accuracy"]
    if has_kappa:
        headers.append("Kappa")

    for c, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = font_header_white
        cell.fill = fill_header
        cell.alignment = center
        cell.border = border_all

    # --- Class data rows (rows 2 .. k+1) ---
    for i in range(k):
        r = i + 2
        # Row label
        cell = ws.cell(row=r, column=1, value=labels[i])
        cell.font = header_font
        cell.alignment = center
        cell.border = border_all

        # Core matrix values
        for j in range(k):
            cell = ws.cell(row=r, column=j + 2, value=float(matrix[i, j]))
            cell.number_format = num_fmt_int
            cell.alignment = center
            cell.border = border_all
            if i == j:
                cell.fill = fill_diag

        # Row total
        cell = ws.cell(row=r, column=k + 2, value=float(row_totals[i]))
        cell.number_format = num_fmt_int
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_total

        # User's Accuracy (P_Accuracy column)
        cell = ws.cell(row=r, column=k + 3, value=round(ua_values[i], 2))
        cell.number_format = num_fmt_pct
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_metric

        # Kappa column (empty for class rows)
        if has_kappa:
            cell = ws.cell(row=r, column=k + 4, value="")
            cell.alignment = center
            cell.border = border_all

    # --- Total row ---
    total_row = k + 2
    cell = ws.cell(row=total_row, column=1, value="Total")
    cell.font = header_font
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_total

    for j in range(k):
        cell = ws.cell(row=total_row, column=j + 2, value=float(col_totals[j]))
        cell.number_format = num_fmt_int
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_total

    cell = ws.cell(row=total_row, column=k + 2, value=float(grand_total))
    cell.number_format = num_fmt_int
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_total
    cell.font = header_font

    # Empty P_Accuracy & Kappa cells in total row
    cell = ws.cell(row=total_row, column=k + 3, value="")
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_total
    if has_kappa:
        cell = ws.cell(row=total_row, column=k + 4, value="")
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_total

    # --- P_Accuracy row (Producer's Accuracy) ---
    pa_row = k + 3
    cell = ws.cell(row=pa_row, column=1, value="P_Accuracy")
    cell.font = header_font
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_metric

    for j in range(k):
        cell = ws.cell(row=pa_row, column=j + 2, value=round(pa_values[j], 2))
        cell.number_format = num_fmt_pct
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_metric

    # Total column in PA row (empty)
    cell = ws.cell(row=pa_row, column=k + 2, value="")
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_metric

    # OA at PA/UA intersection
    cell = ws.cell(row=pa_row, column=k + 3, value=round(oa, 2))
    cell.number_format = num_fmt_pct
    cell.alignment = center
    cell.border = border_all
    cell.fill = fill_metric
    cell.font = header_font

    if has_kappa:
        cell = ws.cell(row=pa_row, column=k + 4, value="")
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_metric

    # --- Kappa row (if available) ---
    if has_kappa:
        kappa_row = k + 4
        cell = ws.cell(row=kappa_row, column=1, value="Kappa")
        cell.font = header_font
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_kappa

        for c in range(2, k + 4):
            cell = ws.cell(row=kappa_row, column=c, value="")
            cell.alignment = center
            cell.border = border_all
            cell.fill = fill_kappa

        # Kappa value in its column
        cell = ws.cell(row=kappa_row, column=k + 4, value=round(result.kappa, 4))
        cell.number_format = "0.0000"
        cell.alignment = center
        cell.border = border_all
        cell.fill = fill_kappa
        cell.font = header_font

    # --- Thick borders around sections ---
    last_col = k + 3 + (1 if has_kappa else 0)
    last_row = (k + 4) if has_kappa else (k + 3)

    # Bottom border on header row
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.border = Border(left=cell.border.left, right=cell.border.right,
                             top=thick, bottom=thick)

    # Right border after class name column
    for r in range(1, last_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.border = Border(left=thick, right=thick,
                             top=cell.border.top, bottom=cell.border.bottom)

    # Thick border separating core from Total column
    for r in range(1, last_row + 1):
        cell = ws.cell(row=r, column=k + 2)
        cell.border = Border(left=thick, right=cell.border.right,
                             top=cell.border.top, bottom=cell.border.bottom)

    # Thick border separating core rows from Total row
    for c in range(1, last_col + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.border = Border(left=cell.border.left, right=cell.border.right,
                             top=thick, bottom=cell.border.bottom)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 16
    for c in range(2, last_col + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        ws.column_dimensions[col_letter].width = 14

    ws.freeze_panes = "B2"

    wb.save(path)


class PerClassMetricsDialog(QDialog):
    """Dialog showing per-class accuracy metrics."""

    def __init__(self, result: ConfusionMatrixResult, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GeoAccuRate \u2014 Per-Class Metrics")
        self.setMinimumSize(650, 400)
        self._build_ui(result)

    def _build_ui(self, result):
        layout = QVBoxLayout(self)

        headers = ["Class", "PA", "PA CI", "UA", "UA CI", "F1"]
        k = len(result.class_labels)
        table = QTableWidget(k, len(headers))
        table.setHorizontalHeaderLabels(headers)

        for i, label in enumerate(result.class_labels):
            name = result.class_names.get(label, str(label))
            pa = result.producers_accuracy.get(label, float("nan"))
            ua = result.users_accuracy.get(label, float("nan"))
            f1 = result.f1_per_class.get(label, float("nan"))
            pa_ci = result.producers_accuracy_ci.get(
                label, (float("nan"), float("nan"))
            )
            ua_ci = result.users_accuracy_ci.get(
                label, (float("nan"), float("nan"))
            )

            def _fmt(v):
                return f"{v:.1%}" if not math.isnan(v) else "\u2014"

            def _fmt_ci(ci):
                lo, hi = ci
                return f"{lo:.1%}\u2013{hi:.1%}" if not math.isnan(lo) else "\u2014"

            values = [
                name,
                _fmt(pa),
                _fmt_ci(pa_ci),
                _fmt(ua),
                _fmt_ci(ua_ci),
                _fmt(f1),
            ]
            for j, text in enumerate(values):
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                table.setItem(i, j, item)

        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table)

        btn_close = QPushButton("Close")
        btn_close.clicked.connect(self.close)
        layout.addWidget(btn_close)
